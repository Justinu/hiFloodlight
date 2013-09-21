#!/usr/bin/python
#coding=utf-8

__author__ = 'justin.bto@gmail.com (Justin Zhou)'

import os
from openpyxl import load_workbook

#定义一个函数读取需导入的包含spotlight的excel

def read_spotlight_dic(import_path):
        
        file_path = import_path
        WorkbookImported = load_workbook(filename = file_path)
        WorksheetImported = WorkbookImported.worksheets[0]
        print WorksheetImported.cell(row = 0, column = 0).value
        
        sheet_column_num = WorksheetImported.get_highest_row()
        sheet_row_num = WorksheetImported.get_highest_row()
        
        #新建一个空的dict来保存读取到的excel表格中的单元格的值
        spotlight_dic = {}
        
        for c in range(sheet_column_num):
            for r in range(sheet_row_num):
                column_tag = WorksheetImported.cell(row = 0, column = c).value
                cell_value = WorksheetImported.cell(row = r, column = c).value
                if (r == 0):
                    spotlight_dic[column_tag] = []
                else:
                    spotlight_dic[column_tag].append(cell_value)
        
        return spotlight_dic
